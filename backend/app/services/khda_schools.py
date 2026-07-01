"""Loads KHDA's 'Dubai Private Schools Open Data' XLSX (ratings + coordinates).

Download 'Dubai's Private Schools Open Data' from
https://web.khda.gov.ae/en/Resources/KHDA-data-statistics (a
DubaiPrivateSchoolsOpenData.xlsx file) and drop it in backend/data/. Each
school row carries lat/long, curriculum, grades, and one DSIB inspection
rating column per academic year; we surface each school's most recent rating.

This adds KHDA ratings on top of the generic OpenStreetMap school POIs, which
have no quality signal.
"""
import glob
import os

import openpyxl

from app.schemas import SchoolNearby
from app.services.dld_csv import DATA_DIR
from app.services.overpass import _haversine_m

# The XLSX has a merged title in row 1; the real header is the next row.
VALID_RATINGS = {"Outstanding", "Very Good", "Good", "Acceptable", "Weak", "Unsatisfactory"}


class _School:
    __slots__ = ("name", "lat", "lon", "curriculum", "grades", "rating", "rating_year")

    def __init__(self, name, lat, lon, curriculum, grades, rating, rating_year):
        self.name = name
        self.lat = lat
        self.lon = lon
        self.curriculum = curriculum
        self.grades = grades
        self.rating = rating
        self.rating_year = rating_year


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Return (header_row_index, {column_name: col_index}) by locating the row
    containing 'School Name'."""
    for ri, row in enumerate(ws.iter_rows(values_only=True, max_row=10)):
        for ci, value in enumerate(row):
            if isinstance(value, str) and value.strip() == "School Name":
                cols = {
                    (v.strip() if isinstance(v, str) else v): i
                    for i, v in enumerate(row)
                    if v is not None
                }
                return ri, cols
    raise ValueError("Could not find the 'School Name' header row in KHDA sheet")


def _rating_columns(header: dict) -> list[tuple[str, int]]:
    """(year_label, col_index) for each DSIB rating column, in sheet order."""
    out = []
    for name, idx in header.items():
        if isinstance(name, str) and "Rating" in name:
            year = name.split("\n")[0].strip()  # '2024/25\nDSIB Rating' -> '2024/25'
            out.append((year, idx))
    return out


def _latest_rating(row: tuple, rating_cols: list[tuple[str, int]]) -> tuple[str | None, str | None]:
    """Most recent valid rating scanning rating columns left→right (oldest→newest)."""
    rating, year = None, None
    for yr, idx in rating_cols:
        value = row[idx] if idx < len(row) else None
        if isinstance(value, str) and value.strip() in VALID_RATINGS:
            rating, year = value.strip(), yr
    return rating, year


def _load() -> list[_School]:
    paths = sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsx")))
    schools: list[_School] = []
    for path in paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = next((s for s in wb.sheetnames if s.lower().startswith("main information")), wb.sheetnames[0])
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        try:
            header_ri, header = _find_header_row(ws)
        except ValueError:
            continue  # not a KHDA schools workbook

        needed = ("School Name", "Latitude", "Longitude")
        if not all(k in header for k in needed):
            continue
        rating_cols = _rating_columns(header)

        for row in rows[header_ri + 1:]:
            name = row[header["School Name"]] if header["School Name"] < len(row) else None
            lat = row[header["Latitude"]] if header["Latitude"] < len(row) else None
            lon = row[header["Longitude"]] if header["Longitude"] < len(row) else None
            if not name or lat is None or lon is None:
                continue
            try:
                lat, lon = float(lat), float(lon)
            except (TypeError, ValueError):
                continue

            rating, year = _latest_rating(row, rating_cols)
            schools.append(
                _School(
                    name=str(name).strip(),
                    lat=lat,
                    lon=lon,
                    curriculum=(str(row[header["Curriculum"]]).strip() if "Curriculum" in header and header["Curriculum"] < len(row) and row[header["Curriculum"]] else None),
                    grades=(str(row[header["Grades"]]).strip() if "Grades" in header and header["Grades"] < len(row) and row[header["Grades"]] else None),
                    rating=rating,
                    rating_year=year,
                )
            )
    return schools


SCHOOLS: list[_School] = _load()


def has_data() -> bool:
    return len(SCHOOLS) > 0


def find_nearby(lat: float, lon: float, radius_m: int = 3000) -> list[SchoolNearby]:
    out: list[SchoolNearby] = []
    for s in SCHOOLS:
        dist = _haversine_m(lat, lon, s.lat, s.lon)
        if dist <= radius_m:
            out.append(
                SchoolNearby(
                    name=s.name,
                    rating=s.rating,
                    rating_year=s.rating_year,
                    curriculum=s.curriculum,
                    grades=s.grades,
                    lat=s.lat,
                    lon=s.lon,
                    distance_m=round(dist, 1),
                )
            )
    return sorted(out, key=lambda s: s.distance_m)
